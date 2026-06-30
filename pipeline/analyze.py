"""
analyze.py — ES task index scroll + individual batch lookup → performance Excel
"""
import logging
import os
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
import urllib3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

urllib3.disable_warnings()
log = logging.getLogger(__name__)

# ── style constants ────────────────────────────────────────────────────────────
_thin       = Side(border_style="thin", color="CCCCCC")
_BORDER     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_BANNER_FILL = PatternFill("solid", fgColor="17365D")
_HDR_FILL    = PatternFill("solid", fgColor="1A6496")
_TOTAL_FILL  = PatternFill("solid", fgColor="EEEEEE")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")

FLAG_COLOR = {
    "HIGH":         "1A7A1A",   # green  >=95%
    "MODERATE":     "E06000",   # orange 70-95%
    "LOW":          "CC0000",   # red    <70%
    "NO TARGET":    "888888",
    "LOW ACTIVITY": "888888",
}

_BATCH = 5000
_WORKERS = 8


# ── ES helpers ─────────────────────────────────────────────────────────────────

def _scroll_all(url, index, query, auth, label):
    """Scroll ES index and return all hits as a list. Use _scroll_batches for large datasets."""
    hits = []
    sid  = None
    try:
        r = requests.post(f"{url}/{index}/_search?scroll=10m",
                          json=query, auth=auth, verify=False, timeout=120)
        r.raise_for_status()
        data  = r.json()
        sid   = data["_scroll_id"]
        batch = data["hits"]["hits"]
        hits.extend(batch)
        total = data["hits"]["total"]["value"]
        log.info(f"  {label}: ~{total:,} docs, fetching ...")
        while batch:
            r = requests.post(f"{url}/_search/scroll",
                              json={"scroll": "10m", "scroll_id": sid},
                              auth=auth, verify=False, timeout=120)
            r.raise_for_status()
            data  = r.json()
            sid   = data["_scroll_id"]
            batch = data["hits"]["hits"]
            hits.extend(batch)
            if len(hits) % 50_000 < len(batch):
                log.info(f"  {label}: {len(hits):,} / ~{total:,} fetched ...")
    finally:
        if sid:
            requests.delete(f"{url}/_search/scroll",
                            json={"scroll_id": sid}, auth=auth, verify=False, timeout=30)
    log.info(f"  {label}: {len(hits):,} docs fetched")
    return hits


def _scroll_batches(url, index, query, auth, label):
    """
    Generator that yields one ES scroll page at a time.
    Never accumulates all docs — safe for millions of records.
    """
    sid = None
    try:
        r = requests.post(f"{url}/{index}/_search?scroll=10m",
                          json=query, auth=auth, verify=False, timeout=120)
        r.raise_for_status()
        data  = r.json()
        sid   = data["_scroll_id"]
        batch = data["hits"]["hits"]
        total = data["hits"]["total"]["value"]
        log.info(f"  {label}: ~{total:,} docs, streaming in batches ...")
        processed = 0
        while batch:
            yield batch
            processed += len(batch)
            if processed % 100_000 < len(batch):
                log.info(f"  {label}: {processed:,} / ~{total:,} streamed ...")
            r = requests.post(f"{url}/_search/scroll",
                              json={"scroll": "10m", "scroll_id": sid},
                              auth=auth, verify=False, timeout=120)
            r.raise_for_status()
            data  = r.json()
            sid   = data["_scroll_id"]
            batch = data["hits"]["hits"]
    finally:
        if sid:
            requests.delete(f"{url}/_search/scroll",
                            json={"scroll_id": sid}, auth=auth, verify=False, timeout=30)


def _build_campaign_filters(cfg):
    """
    Build ES filter clauses that identify this specific campaign in the task index.

    Only applied when task_campaign_filter=TRUE (opt-in).
    - Nigeria SMC states: FALSE — date range alone isolates the campaign.
    - AZM/non-admin (multiple project types share same tenant+date): TRUE.

    Priority (first match wins):
      1. is_admin_console=TRUE  → campaignNumber  (Nigeria admin, Chad admin)
      2. project_type_id set    → projectTypeId   (Togo non-admin)
      3. project_type set       → projectType     (AZM Nigeria/Congo)
      Appends cycleIndex filter if cycle_index is set (regardless of above).
    """
    if not cfg.get("task_campaign_filter", False):
        return []

    filters = []
    if cfg.get("is_admin_console") and cfg.get("campaign_number"):
        filters.append({"term": {"Data.campaignNumber.keyword": cfg["campaign_number"]}})
    elif cfg.get("project_type_id"):
        filters.append({"term": {"Data.projectTypeId.keyword": cfg["project_type_id"]}})
    elif cfg.get("project_type"):
        filters.append({"term": {"Data.projectType.keyword": cfg["project_type"]}})

    if cfg.get("cycle_index"):
        filters.append({"term": {"Data.additionalDetails.cycleIndex.keyword": cfg["cycle_index"]}})

    return filters


def _fetch_task_docs(cfg):
    _source = [
        "Data.boundaryHierarchy", "Data.age", "Data.individualId",
        "Data.quantity", "Data.administrationStatus",
        "Data.additionalDetails",
    ]
    date_field  = cfg.get("task_date_field", "taskDates")
    date_filter = {"range": {f"Data.{date_field}": {"gte": cfg["GTE"], "lte": cfg["LTE"]}}}
    campaign_f  = _build_campaign_filters(cfg)

    # Base filters shared by both queries
    base = [date_filter] + campaign_f

    # Query 1: treatment records
    treatment_filters = base + [
        {"terms": {"Data.administrationStatus.keyword": [
            "ADMINISTRATION_SUCCESS", "VISITED",
        ]}},
    ]
    # doseIndex=1 filter: opt-in only (dose_index_filter=TRUE).
    # FALSE for all Nigeria SMC states and AZM — extraction scripts confirm doseIndex unused.
    if cfg.get("dose_index_filter"):
        treatment_filters.append({"term": {"Data.additionalDetails.doseIndex.keyword": "1"}})

    q_treatment = {
        "size": _BATCH,
        "query": {"bool": {"filter": treatment_filters}},
        "_source": _source,
    }

    # Query 2: non-admin records — no doseIndex filter (absent/refused/etc. have no doseIndex)
    q_nonadmin = {
        "size": _BATCH,
        "query": {"bool": {"filter": base + [
            {"terms": {"Data.administrationStatus.keyword": [
                "BENEFICIARY_INELIGIBLE", "INELIGIBLE",
                "BENEFICIARY_REFERRED", "BENEFICIARY_DIED",
                "BENEFICIARY_ABSENT", "BENEFICIARY_MIGRATED",
                "BENEFICIARY_REFUSED",
            ]}},
        ]}},
        "_source": _source,
    }

    hits  = _scroll_all(cfg["es_url"], cfg["ES_INDEX_TASK"], q_treatment, cfg["es_auth"], "task-treatment")
    hits += _scroll_all(cfg["es_url"], cfg["ES_INDEX_TASK"], q_nonadmin,  cfg["es_auth"], "task-nonadmin")
    return hits


def _fetch_individual_names(cfg, ind_ids):
    """Batch-fetch child names from individual index. Returns {clientReferenceId: name}."""
    if not ind_ids:
        return {}
    batches = [ind_ids[i:i + _BATCH] for i in range(0, len(ind_ids), _BATCH)]
    name_map = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"terms": {"clientReferenceId.keyword": batch}},
            "_source": ["clientReferenceId", "name"],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_IND']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        result = {}
        for h in r.json()["hits"]["hits"]:
            src = h["_source"]
            cid = src.get("clientReferenceId", "")
            n   = src.get("name", {})
            given  = n.get("givenName")  or ""
            family = n.get("familyName") or ""
            result[cid] = f"{given} {family}".strip()
        return result

    log.info(f"  individual lookup: {len(ind_ids):,} IDs in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                name_map.update(f.result())
            except Exception as e:
                log.warning(f"  individual batch error: {e}")
    log.info(f"  individual lookup: {len(name_map):,} names resolved")
    return name_map


def _fetch_hh_member_map(cfg, ind_ids):
    """
    Step 4 (Togo pattern): for each individual ID find their household clientReferenceId.
    Queries household-member-index by Data.householdMember.individualClientReferenceId.
    Returns {ind_clref_id: hh_clref_id}.
    """
    if not ind_ids:
        return {}
    batches    = [ind_ids[i:i + _BATCH] for i in range(0, len(ind_ids), _BATCH)]
    member_map = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"terms": {
                "Data.householdMember.individualClientReferenceId.keyword": batch
            }},
            "_source": [
                "Data.householdMember.individualClientReferenceId",
                "Data.householdMember.householdClientReferenceId",
            ],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_HH_MEMBER']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        result = {}
        for h in r.json()["hits"]["hits"]:
            src    = h["_source"].get("Data", {}).get("householdMember", {})
            ind_id = src.get("individualClientReferenceId", "")
            hh_id  = src.get("householdClientReferenceId", "")
            if ind_id and hh_id:
                result[ind_id] = hh_id
        return result

    log.info(f"  HH member map: {len(ind_ids):,} individuals in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                member_map.update(f.result())
            except Exception as e:
                log.warning(f"  HH member batch error: {e}")
    log.info(f"  HH member map: {len(member_map):,} memberships resolved")
    return member_map


def _fetch_hh_head_map(cfg, hh_clref_ids):
    """
    Step 5 (Togo pattern): for each household ID find the head-of-household's individual ID.
    Queries household-member-index by Data.householdMember.householdClientReferenceId
    filtered to isHeadOfHousehold=True.
    Returns {hh_clref_id: head_individual_clref_id}.
    """
    if not hh_clref_ids:
        return {}
    batches  = [hh_clref_ids[i:i + _BATCH] for i in range(0, len(hh_clref_ids), _BATCH)]
    head_map = {}

    def _one_batch(batch):
        q = {
            "size": _BATCH,
            "query": {"bool": {"must": [
                {"terms": {"Data.householdMember.householdClientReferenceId.keyword": batch}},
                {"term":  {"Data.householdMember.isHeadOfHousehold": True}},
            ]}},
            "_source": [
                "Data.householdMember.householdClientReferenceId",
                "Data.householdMember.individualClientReferenceId",
            ],
        }
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_HH_MEMBER']}/_search",
            json=q, auth=cfg["es_auth"], verify=False, timeout=60,
        )
        r.raise_for_status()
        result = {}
        for h in r.json()["hits"]["hits"]:
            src    = h["_source"].get("Data", {}).get("householdMember", {})
            hh_id  = src.get("householdClientReferenceId", "")
            ind_id = src.get("individualClientReferenceId", "")
            if hh_id and ind_id:
                result[hh_id] = ind_id
        return result

    log.info(f"  HH head map: {len(hh_clref_ids):,} households in {len(batches)} batches ...")
    with ThreadPoolExecutor(max_workers=_WORKERS) as ex:
        futures = {ex.submit(_one_batch, b): b for b in batches}
        for f in as_completed(futures):
            try:
                head_map.update(f.result())
            except Exception as e:
                log.warning(f"  HH head batch error: {e}")
    log.info(f"  HH head map: {len(head_map):,} heads resolved")
    return head_map


# ── aggregation ────────────────────────────────────────────────────────────────

def _load_targets(cfg):
    csv_path = cfg["target_csv"]
    if not csv_path:
        log.warning("target_csv not set — all targets = 0")
        return {}

    # Google Sheets URL: extract sheet_id and optional gid
    if csv_path.startswith("https://docs.google.com/spreadsheets/"):
        import re
        import gspread
        from google.oauth2.service_account import Credentials
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), ".env"))

        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", csv_path)
        if not m:
            log.warning(f"Could not parse sheet ID from target_csv URL: {csv_path}")
            return {}
        sheet_id = m.group(1)
        gid_m = re.search(r"[#&]gid=(\d+)", csv_path)

        creds_path = os.getenv("GOOGLE_CREDENTIALS_PATH")
        creds = Credentials.from_service_account_file(
            creds_path,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        client = gspread.Client(auth=creds)
        spreadsheet = client.open_by_key(sheet_id)
        if gid_m:
            ws = next(
                (s for s in spreadsheet.worksheets() if str(s.id) == gid_m.group(1)),
                spreadsheet.sheet1,
            )
        else:
            ws = spreadsheet.sheet1

        records = ws.get_all_records()
        df = pd.DataFrame(records)
        log.info(f"target_csv loaded from Google Sheet: {sheet_id} ({len(df)} rows)")
    elif not os.path.exists(csv_path):
        log.warning(f"target_csv not found: {csv_path} — all targets = 0")
        return {}
    else:
        df = pd.read_csv(csv_path)
    col_name = "facility_name" if "facility_name" in df.columns else df.columns[0]
    col_tgt  = "individual_target" if "individual_target" in df.columns else df.columns[1]
    tmap = {}
    for _, row in df.iterrows():
        name = str(row[col_name]).strip().lower()
        try:
            itgt = float(row[col_tgt])
        except Exception:
            itgt = 0.0
        daily = round(itgt / cfg["campaign_days"]) if itgt > 0 else 0
        tmap[name] = {"4day": int(itgt), "daily": daily}
    log.info(f"  targets loaded: {len(tmap)} facilities")
    return tmap


def _band(treat_rate, daily_target, records, drug_type):
    # LOW ACTIVITY: <10 records regardless of coverage rate
    if records < 10:
        return "LOW ACTIVITY"
    if daily_target == 0:
        return "NO TARGET"
    # Coverage bands: HIGH >=95% | MODERATE 70-95% | LOW <70%
    if treat_rate >= 95:
        return "HIGH"
    if treat_rate >= 70:
        return "MODERATE"
    return "LOW"


def _aggregate_batch(task_hits, name_map, hh_name_map, fac_data, cfg):
    """
    Process one batch of task hits, updating the running fac_data dict in-place.
    Uses integer hash keys for dedup to keep memory ~7x lower than storing full tuples.
    """
    drug_type = cfg["drug_type"]
    min_age   = 3 if drug_type == "SPAQ" else 1

    for h in task_hits:
        doc  = h["_source"]["Data"]
        bh   = doc.get("boundaryHierarchy") or {}
        lga  = str(bh.get("lga",            "") or "").strip()
        ward = str(bh.get("ward",           "") or "").strip()
        fac  = str(bh.get("healthFacility", "") or "").strip()
        if not fac:
            continue

        ind_id     = doc.get("individualId", "") or ""
        adm        = doc.get("administrationStatus", "") or ""
        add        = doc.get("additionalDetails") or {}
        lat        = add.get("latitude")
        lon        = add.get("longitude")
        del_com    = str(add.get("deliveryComments") or doc.get("deliveryComments") or "").strip()
        hh_head    = hh_name_map.get(ind_id, "")
        child_name = name_map.get(ind_id, "") if ind_id else ""

        age_raw = doc.get("age")
        try:
            age = int(float(age_raw)) if age_raw not in (None, "") else None
        except Exception:
            age = None

        qty_raw = doc.get("quantity")
        try:
            qty = int(float(qty_raw)) if qty_raw not in (None, "") else 0
        except Exception:
            qty = 0

        inelig_status = adm in ("BENEFICIARY_INELIGIBLE", "INELIGIBLE")
        ref_status    = adm == "BENEFICIARY_REFERRED"

        if fac not in fac_data:
            fac_data[fac] = dict(
                lga=lga, fac=fac, records=0, treated=0,
                drug1=0, drug2=0, absent=0, refused=0,
                ineligible=0, referred=0, died=0, migrated=0,
                redose=0, age_over59=0, age_zero=0,
                missing_hh=0, missing_child=0, missing_gender=0, duplicates=0,
                delivery_comments=0, missing_lat_lon=0,
                wards=set(), seen_keys=set(),
            )

        m = fac_data[fac]
        m["records"] += 1

        if ward:
            m["wards"].add(ward)
        if (lat is None or lat == "") or (lon is None or lon == ""):
            m["missing_lat_lon"] += 1
        if del_com:
            m["delivery_comments"] += 1
        if not child_name:
            m["missing_child"] += 1
        if not hh_head:
            m["missing_hh"] += 1
        if not add.get("gender"):
            m["missing_gender"] += 1

        if age is None:
            pass
        elif age == 0:
            m["age_zero"] += 1
        elif age > 59:
            m["age_over59"] += 1

        if   adm == "BENEFICIARY_ABSENT":                         m["absent"]    += 1
        elif adm == "BENEFICIARY_REFUSED":                        m["refused"]   += 1
        elif adm in ("BENEFICIARY_INELIGIBLE", "INELIGIBLE"):    m["ineligible"] += 1
        elif adm == "BENEFICIARY_REFERRED":                       m["referred"]  += 1
        elif adm == "BENEFICIARY_DIED":                           m["died"]      += 1
        elif adm == "BENEFICIARY_MIGRATED":                       m["migrated"]  += 1

        if adm == "VISITED" and qty >= 1:
            m["redose"] += 1

        is_treated = (
            age is not None and min_age <= age <= 59
            and not inelig_status and not ref_status
            and qty >= 1 and adm != "VISITED"
        )
        if is_treated:
            m["treated"] += 1
            if age is not None and age <= 11:
                m["drug2"] += 1
            else:
                m["drug1"] += 1

        # dedup: store hash(tuple) — uses ~28 bytes vs ~200 bytes for full tuple
        if child_name and age is not None:
            key_hash = hash((hh_head.lower(), child_name.lower(), ward.lower(), age))
            if key_hash in m["seen_keys"]:
                m["duplicates"] += 1
            else:
                m["seen_keys"].add(key_hash)


def _finalize_fac_data(fac_data, target_map, cfg):
    """Convert running fac_data dict into the final sorted results list."""
    results = []
    for fac, m in sorted(fac_data.items(), key=lambda x: (x[1]["lga"], x[0])):
        tgt_entry   = target_map.get(fac.lower(), {"4day": 0, "daily": 0})
        daily_tgt   = tgt_entry["daily"]
        records     = m["records"]
        treated     = m["treated"]
        not_treated = records - treated

        if daily_tgt > 0:
            rate = treated / daily_tgt * 100
            cov  = f"{rate:.1f}%"
        else:
            rate = 0.0
            cov  = "NO TARGET"

        flag = _band(rate, daily_tgt, records, cfg["drug_type"])

        results.append({
            "lga":               m["lga"],
            "fac":               fac,
            "daily_target":      daily_tgt,
            "records":           records,
            "treated":           treated,
            "not_treated":       not_treated,
            "drug1":             m["drug1"],
            "drug2":             m["drug2"],
            "coverage":          cov,
            "status":            flag,
            "absent":            m["absent"],
            "refused":           m["refused"],
            "ineligible":        m["ineligible"],
            "referred":          m["referred"],
            "died":              m["died"],
            "migrated":          m["migrated"],
            "redose":            m["redose"],
            "age_over59":        m["age_over59"],
            "age_zero":          m["age_zero"],
            "missing_hh":        m["missing_hh"],
            "missing_child":     m["missing_child"],
            "missing_gender":    m["missing_gender"],
            "duplicates":        m["duplicates"],
            "delivery_comments": m["delivery_comments"],
            "wards":             ", ".join(sorted(m["wards"])),
            "rate":              rate,
        })
    return results


def _aggregate(task_hits, name_map, hh_name_map, target_map, cfg):
    drug_type    = cfg["drug_type"]
    campaign_days = cfg["campaign_days"]

    # facility key → metrics dict
    fac_data = {}

    for h in task_hits:
        doc  = h["_source"]["Data"]
        bh   = doc.get("boundaryHierarchy") or {}
        lga  = str(bh.get("lga",           "") or "").strip()
        ward = str(bh.get("ward",          "") or "").strip()
        fac  = str(bh.get("healthFacility","") or "").strip()
        if not fac:
            continue

        ind_id  = doc.get("individualId", "") or ""
        adm     = doc.get("administrationStatus", "") or ""
        add     = doc.get("additionalDetails") or {}
        # lat/lon live in additionalDetails (Togo pattern)
        lat     = add.get("latitude")
        lon     = add.get("longitude")
        del_com = str(add.get("deliveryComments") or doc.get("deliveryComments") or "").strip()
        # hh_head keyed by child ind_id (built in run() via household-member index)
        hh_head = hh_name_map.get(ind_id, "")
        # child name always from individual-index lookup (Togo pattern)
        child_name = name_map.get(ind_id, "") if ind_id else ""

        age_raw = doc.get("age")
        try:
            age = int(float(age_raw)) if age_raw not in (None, "") else None
        except Exception:
            age = None

        qty_raw = doc.get("quantity")
        try:
            qty = int(float(qty_raw)) if qty_raw not in (None, "") else 0
        except Exception:
            qty = 0

        inelig_status = adm in ("BENEFICIARY_INELIGIBLE", "INELIGIBLE")
        ref_status    = adm == "BENEFICIARY_REFERRED"

        if fac not in fac_data:
            fac_data[fac] = dict(
                lga=lga, fac=fac, records=0, treated=0,
                drug1=0, drug2=0, absent=0, refused=0,
                ineligible=0, referred=0, died=0, migrated=0,
                redose=0, age_over59=0, age_zero=0,
                missing_hh=0, missing_child=0, missing_gender=0, duplicates=0,
                delivery_comments=0, missing_lat_lon=0,
                wards=set(), seen_keys=set(),
            )

        m = fac_data[fac]
        m["records"] += 1

        if ward:
            m["wards"].add(ward)
        if (lat is None or lat == "") or (lon is None or lon == ""):
            m["missing_lat_lon"] += 1
        if del_com:
            m["delivery_comments"] += 1
        if not child_name:
            m["missing_child"] += 1
        if not hh_head:
            m["missing_hh"] += 1
        if not add.get("gender"):
            m["missing_gender"] += 1

        if age is None:
            pass
        elif age == 0:
            m["age_zero"] += 1
        elif age > 59:
            m["age_over59"] += 1

        # non-admin
        if   adm == "BENEFICIARY_ABSENT":    m["absent"]    += 1
        elif adm == "BENEFICIARY_REFUSED":   m["refused"]   += 1
        elif adm in ("BENEFICIARY_INELIGIBLE","INELIGIBLE"): m["ineligible"] += 1
        elif adm == "BENEFICIARY_REFERRED":  m["referred"]  += 1
        elif adm == "BENEFICIARY_DIED":      m["died"]      += 1
        elif adm == "BENEFICIARY_MIGRATED":  m["migrated"]  += 1

        # redose
        if adm == "VISITED" and qty >= 1:
            m["redose"] += 1

        # treatment
        if drug_type == "SPAQ":
            min_age = 3
        else:
            min_age = 1
        is_treated = (
            age is not None and min_age <= age <= 59
            and not inelig_status and not ref_status
            and qty >= 1
            and adm != "VISITED"
        )
        if is_treated:
            m["treated"] += 1
            if age is not None and age <= 11:
                m["drug2"] += 1   # SPAQ1 (3-11m) / AZM 1-11m
            else:
                m["drug1"] += 1   # SPAQ2 (12-59m) / AZM 12-59m

        # dedup: (hh_head, child_name, ward, age)
        if child_name and age is not None:
            key = (hh_head.lower(), child_name.lower(), ward.lower(), age)
            if key in m["seen_keys"]:
                m["duplicates"] += 1
            else:
                m["seen_keys"].add(key)

    # resolve daily targets
    results = []
    for fac, m in sorted(fac_data.items(), key=lambda x: (x[1]["lga"], x[0])):
        tgt_entry   = target_map.get(fac.lower(), {"4day": 0, "daily": 0})
        daily_tgt   = tgt_entry["daily"]
        four_day    = tgt_entry["4day"]
        records     = m["records"]
        treated     = m["treated"]
        not_treated = records - treated

        if daily_tgt > 0:
            rate = treated / daily_tgt * 100
            cov  = f"{rate:.1f}%"
        else:
            rate = 0.0
            cov  = "NO TARGET"

        flag = _band(rate, daily_tgt, records, cfg["drug_type"])

        results.append({
            "lga":              m["lga"],
            "fac":              fac,
            "daily_target":     daily_tgt,
            "records":          records,
            "treated":          treated,
            "not_treated":      not_treated,
            "drug1":            m["drug1"],
            "drug2":            m["drug2"],
            "coverage":         cov,
            "status":           flag,
            "absent":           m["absent"],
            "refused":          m["refused"],
            "ineligible":       m["ineligible"],
            "referred":         m["referred"],
            "died":             m["died"],
            "migrated":         m["migrated"],
            "redose":           m["redose"],
            "age_over59":       m["age_over59"],
            "age_zero":         m["age_zero"],
            "missing_hh":       m["missing_hh"],
            "missing_child":    m["missing_child"],
            "missing_gender":   m["missing_gender"],
            "duplicates":       m["duplicates"],
            "delivery_comments":m["delivery_comments"],
            "wards":            ", ".join(sorted(m["wards"])),
            "rate":             rate,
        })

    return results


# ── Excel writing ──────────────────────────────────────────────────────────────

def _col_headers(drug_type):
    d1 = "SPAQ2 (12-59m)" if drug_type == "SPAQ" else "AZM 12-59m"
    d2 = "SPAQ1 (3-11m)"  if drug_type == "SPAQ" else "AZM 1-11m"
    return [
        "#", "LGA", "Health Facility", "Daily Target", "Records", "Treated",
        "Not Treated", d1, d2, "Coverage %", "Status",
        "Absent", "Refused", "Ineligible", "Referred", "Died", "Migrated",
        "Redose", "Age>59", "Age=0", "Missing HH", "Missing Child", "Missing Gender",
        "Duplicates", "Delivery Comments", "Wards",
    ]


def _row_values(r, idx):
    return [
        idx, r["lga"], r["fac"], r["daily_target"], r["records"], r["treated"],
        r["not_treated"], r["drug1"], r["drug2"], r["coverage"], r["status"],
        r["absent"], r["refused"], r["ineligible"], r["referred"], r["died"],
        r["migrated"], r["redose"], r["age_over59"], r["age_zero"],
        r["missing_hh"], r["missing_child"], r["missing_gender"], r["duplicates"],
        r["delivery_comments"], r["wards"],
    ]


def _totals(rows, drug_type):
    def s(k): return sum(r[k] for r in rows)
    total_tgt  = s("daily_target")
    total_rec  = s("records")
    total_tr   = s("treated")
    cov = f"{total_tr/total_tgt*100:.1f}%" if total_tgt else "N/A"
    d1 = "SPAQ2 (12-59m)" if drug_type == "SPAQ" else "AZM 12-59m"
    d2 = "SPAQ1 (3-11m)"  if drug_type == "SPAQ" else "AZM 1-11m"
    return {
        "lga": "", "fac": "GRAND TOTAL",
        "daily_target": total_tgt, "records": total_rec, "treated": total_tr,
        "not_treated": s("not_treated"), "drug1": s("drug1"), "drug2": s("drug2"),
        "coverage": cov, "status": "",
        "absent": s("absent"), "refused": s("refused"), "ineligible": s("ineligible"),
        "referred": s("referred"), "died": s("died"), "migrated": s("migrated"),
        "redose": s("redose"), "age_over59": s("age_over59"), "age_zero": s("age_zero"),
        "missing_hh": s("missing_hh"), "missing_child": s("missing_child"),
        "missing_gender": s("missing_gender"),
        "duplicates": s("duplicates"), "delivery_comments": s("delivery_comments"),
        "wards": "", "rate": 0,
    }


def _style_cell(cell, fill=None, bold=False, color=None, align="center", size=9):
    cell.border = _BORDER
    if fill:
        cell.fill = fill
    font_kwargs = {"bold": bold, "size": size, "name": "Calibri"}
    if color:
        from openpyxl.styles import colors
        font_kwargs["color"] = color
    cell.font   = Font(**font_kwargs)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)


def _write_tab(ws, rows, headers, drug_type, cfg, banner_text):
    ncols = len(headers)
    last_col = get_column_letter(ncols)

    # row 1: banner
    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = _BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    # row 2: headers
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        _style_cell(cell, fill=_HDR_FILL, bold=True, align="center", color="FFFFFF")

    # data rows
    for ri, r in enumerate(rows, 1):
        vals = _row_values(r, ri)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            _style_cell(cell, fill=_WHITE_FILL, align="center")
            # color Status and Coverage % cols
            if ci == 10:  # Coverage %
                pass
            if ci == 11:  # Status
                flag_col = FLAG_COLOR.get(str(val), "000000")
                cell.font = Font(bold=True, color=flag_col, size=9, name="Calibri")

    # totals row
    if rows:
        tot = _totals(rows, drug_type)
        tot_row = len(rows) + 3
        tot_vals = _row_values(tot, "")
        for ci, val in enumerate(tot_vals, 1):
            cell = ws.cell(row=tot_row, column=ci, value=val)
            _style_cell(cell, fill=_TOTAL_FILL, bold=True, align="center")

    # freeze, filter, col widths
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    for ci in range(4, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 12
    ws.column_dimensions[get_column_letter(ncols)].width = 30  # Wards


# ── public entry point ─────────────────────────────────────────────────────────

BANDS = ["LOW", "MODERATE", "HIGH", "NO TARGET", "LOW ACTIVITY"]


def run(cfg):
    log.info(f"[analyze] {cfg['state_name']} Day {cfg['DAY']} — streaming task docs ...")

    target_map = _load_targets(cfg)
    fac_data   = {}

    _source    = [
        "Data.boundaryHierarchy", "Data.age", "Data.individualId",
        "Data.quantity", "Data.administrationStatus", "Data.additionalDetails",
    ]
    date_field  = cfg.get("task_date_field", "taskDates")
    date_filter = {"range": {f"Data.{date_field}": {"gte": cfg["GTE"], "lte": cfg["LTE"]}}}
    campaign_f  = _build_campaign_filters(cfg)
    base        = [date_filter] + campaign_f

    treatment_filters = base + [
        {"terms": {"Data.administrationStatus.keyword": ["ADMINISTRATION_SUCCESS", "VISITED"]}},
    ]
    if cfg.get("dose_index_filter"):
        treatment_filters.append({"term": {"Data.additionalDetails.doseIndex.keyword": "1"}})

    nonadmin_filters = base + [
        {"terms": {"Data.administrationStatus.keyword": [
            "BENEFICIARY_INELIGIBLE", "INELIGIBLE", "BENEFICIARY_REFERRED",
            "BENEFICIARY_DIED", "BENEFICIARY_ABSENT", "BENEFICIARY_MIGRATED",
            "BENEFICIARY_REFUSED",
        ]}},
    ]

    queries = [
        ("task-treatment", {"size": _BATCH, "query": {"bool": {"filter": treatment_filters}}, "_source": _source}),
        ("task-nonadmin",  {"size": _BATCH, "query": {"bool": {"filter": nonadmin_filters}},  "_source": _source}),
    ]

    total_processed = 0
    for label, query in queries:
        for batch in _scroll_batches(cfg["es_url"], cfg["ES_INDEX_TASK"], query, cfg["es_auth"], label):
            # Fetch individual names only for this batch — never accumulate all IDs
            batch_ind_ids = list({
                h["_source"]["Data"].get("individualId", "")
                for h in batch
                if h["_source"]["Data"].get("individualId", "")
            })
            name_map = _fetch_individual_names(cfg, batch_ind_ids)
            _aggregate_batch(batch, name_map, {}, fac_data, cfg)
            total_processed += len(batch)

    log.info(f"[analyze] {total_processed:,} records processed across {len(fac_data)} facilities")
    rows = _finalize_fac_data(fac_data, target_map, cfg)

    drug_type = cfg["drug_type"]
    headers   = _col_headers(drug_type)

    total_target = sum(r["daily_target"] for r in rows)
    total_daily  = total_target
    banner_text  = (
        f"Campaign Target: {total_target * cfg['campaign_days']:,}  |  "
        f"Daily Target (Day {cfg['DAY']}): {total_daily:,}"
    )

    wb = Workbook()
    wb.remove(wb.active)

    # ALL FACILITIES tab
    ws_all = wb.create_sheet("ALL FACILITIES")
    _write_tab(ws_all, rows, headers, drug_type, cfg, banner_text)

    # Per-band tabs
    for band in BANDS:
        band_rows = [r for r in rows if r["status"] == band]
        ws = wb.create_sheet(band)
        _write_tab(ws, band_rows, headers, drug_type, cfg, banner_text)

    out = cfg["perf_xlsx"]
    wb.save(out)
    log.info(f"[analyze] saved -> {out}  ({len(rows)} facilities)")
    return out
