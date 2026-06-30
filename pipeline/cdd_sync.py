"""
cdd_sync.py — ES staff index scroll + composite agg → CDD sync Excel

Two variants:
  is_admin_console=TRUE  (Nigeria): filter staff by campaignNumber, sync by syncedUserName
  is_admin_console=FALSE (Chad/Togo): filter staff by projectTypeId, sync by syncedUserId
"""
import logging
from collections import defaultdict
from datetime import date, timedelta, datetime, timezone

import pandas as pd
import requests
import urllib3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

urllib3.disable_warnings()
log = logging.getLogger(__name__)

_HDR_FILL    = PatternFill("solid", fgColor="003366")
_NEVER_FILL  = PatternFill("solid", fgColor="FFD7D7")
_LOW_FILL    = PatternFill("solid", fgColor="FFE0B3")
_TOTAL_FILL  = PatternFill("solid", fgColor="D6E4F0")
_thin        = Side(border_style="thin", color="CCCCCC")
_BORDER      = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── ES helpers ─────────────────────────────────────────────────────────────────

def _scroll_all(url, index, query, auth, label):
    hits = []
    sid  = None
    try:
        r = requests.post(f"{url}/{index}/_search?scroll=10m",
                          json=query, auth=auth, verify=False, timeout=120)
        r.raise_for_status()
        data  = r.json()
        sid   = data["_scroll_id"]
        batch = data["hits"]["hits"]
        hits.extend(batch)
        total = data["hits"]["total"]["value"]
        log.info(f"  {label}: ~{total:,} docs ...")
        while batch:
            r = requests.post(f"{url}/_search/scroll",
                              json={"scroll": "10m", "scroll_id": sid},
                              auth=auth, verify=False, timeout=120)
            r.raise_for_status()
            data  = r.json()
            sid   = data["_scroll_id"]
            batch = data["hits"]["hits"]
            hits.extend(batch)
            if len(hits) % 50_000 < len(batch):
                log.info(f"  {label}: {len(hits):,} / ~{total:,} fetched ...")
    finally:
        if sid:
            requests.delete(f"{url}/_search/scroll",
                            json={"scroll_id": sid}, auth=auth, verify=False, timeout=30)
    log.info(f"  {label}: {len(hits):,} docs")
    return hits


def _composite_agg(url, index, query_must, agg_sources, auth):
    """Paginate through a composite aggregation. Returns list of buckets."""
    buckets = []
    after   = None
    while True:
        agg_body = {"size": 1000, "sources": agg_sources}
        if after:
            agg_body["after"] = after
        q = {
            "size": 0,
            "query": {"bool": {"must": query_must}},
            "aggs": {"combo": {"composite": agg_body}},
        }
        r = requests.post(f"{url}/{index}/_search",
                          json=q, auth=auth, verify=False, timeout=120)
        r.raise_for_status()
        data   = r.json()
        page   = data["aggregations"]["combo"]["buckets"]
        buckets.extend(page)
        after  = data["aggregations"]["combo"].get("after_key")
        if not after:
            break
    return buckets


def _ts_to_date(val):
    if isinstance(val, int):
        return datetime.fromtimestamp(val / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
    return str(val)[:10]


# ── Variant A: admin console = TRUE (Nigeria, campaignNumber) ──────────────────

def _load_staff_admin(cfg):
    query = {
        "size": 5000,
        "_source": [
            "Data.userId", "Data.userName", "Data.nameOfUser",
            "Data.boundaryHierarchy.healthFacility",
            "Data.boundaryHierarchy.lga",
        ],
        "query": {"bool": {"must": [
            {"term": {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term": {"Data.role.keyword": "DISTRIBUTOR"}},
        ]}},
    }
    hits = _scroll_all(cfg["es_url"], cfg["ES_INDEX_STAFF"], query,
                       cfg["es_auth"], "staff (admin)")
    cdds       = {}   # lower(uname) -> info
    uname_list = []
    for h in hits:
        d     = h["_source"]["Data"]
        uname = (d.get("userName", "") or d.get("nameOfUser", "")).strip()
        key   = uname.lower()
        if not key or key in cdds:
            continue
        bh = d.get("boundaryHierarchy") or {}
        cdds[key] = {
            "user_id":  d.get("userId", "").strip(),
            "username": uname,
            "facility": bh.get("healthFacility", ""),
            "lga":      bh.get("lga", ""),
        }
        uname_list.append(uname)
    log.info(f"  staff (admin): {len(cdds):,} unique CDDs")
    return cdds, uname_list


def _load_syncs_admin(cfg, uname_list):
    must = [
        {"terms": {"Data.taskDates":              cfg["CAMPAIGN_DATES"]}},
        {"terms": {"Data.syncedUserName.keyword": uname_list}},
        {"term":  {"Data.role.keyword":           "DISTRIBUTOR"}},
        {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
    ]
    sources = [
        {"uname": {"terms": {"field": "Data.syncedUserName.keyword"}}},
        {"date":  {"terms": {"field": "Data.taskDates"}}},
    ]
    buckets = _composite_agg(cfg["es_url"], cfg["ES_INDEX_SYNC"], must, sources,
                              cfg["es_auth"])
    log.info(f"  sync agg (admin): {len(buckets):,} (uname, date) pairs")
    user_dates = defaultdict(set)
    for b in buckets:
        uname = b["key"]["uname"].lower()
        dt    = _ts_to_date(b["key"]["date"])
        if dt in cfg["CAMPAIGN_DATES"]:
            user_dates[uname].add(dt)
    return user_dates


# ── Variant B: admin console = FALSE (Chad/Togo, projectTypeId) ───────────────

def _load_staff_project(cfg):
    query = {
        "size": 5000,
        "_source": [
            "Data.userId", "Data.userName", "Data.nameOfUser",
            "Data.boundaryHierarchy",
        ],
        "query": {"bool": {"must": [
            {"term": {"Data.projectTypeId.keyword": cfg["project_type_id"]}},
            {"term": {"Data.role.keyword":          "DISTRIBUTOR"}},
            {"term": {"Data.isDeleted":             False}},
        ]}},
    }
    hits = _scroll_all(cfg["es_url"], cfg["ES_INDEX_STAFF"], query,
                       cfg["es_auth"], "staff (project)")
    cdds     = {}   # lower(uid) -> info
    uid_list = []
    for h in hits:
        d   = h["_source"]["Data"]
        uid = (d.get("userId", "") or "").strip().lower()
        if not uid or uid in cdds:
            continue
        bh = d.get("boundaryHierarchy") or {}
        cdds[uid] = {
            "user_id":    d.get("userId", "").strip(),
            "username":   (d.get("userName", "") or d.get("nameOfUser", "")).strip(),
            "province":   bh.get("province",    ""),
            "district":   bh.get("district",    ""),
            "health_area":bh.get("healthArea",  ""),
            "facility":   bh.get("healthFacility", ""),
            "lga":        bh.get("lga", ""),
        }
        uid_list.append(d.get("userId", "").strip())
    log.info(f"  staff (project): {len(cdds):,} unique CDDs")
    return cdds, uid_list


def _load_syncs_project(cfg, uid_list):
    must = [
        {"terms": {"Data.taskDates":            cfg["CAMPAIGN_DATES"]}},
        {"terms": {"Data.syncedUserId.keyword": uid_list}},
    ]
    sources = [
        {"uid":  {"terms": {"field": "Data.syncedUserId.keyword"}}},
        {"date": {"terms": {"field": "Data.taskDates"}}},
    ]
    buckets = _composite_agg(cfg["es_url"], cfg["ES_INDEX_SYNC"], must, sources,
                              cfg["es_auth"])
    log.info(f"  sync agg (project): {len(buckets):,} (uid, date) pairs")
    user_dates = defaultdict(set)
    for b in buckets:
        uid = b["key"]["uid"].lower()
        dt  = _ts_to_date(b["key"]["date"])
        if dt in cfg["CAMPAIGN_DATES"]:
            user_dates[uid].add(dt)
    return user_dates


# ── time-cutoff sync count ────────────────────────────────────────────────────

def _count_synced_by_cutoff(cfg, cutoff_hour, cutoff_min=0):
    """
    Count unique CDDs who synced TODAY before cutoff_hour:cutoff_min UTC.
    Uses Data.createdTime (epoch ms) as the server receipt timestamp.
    Returns int count, or None if query fails.
    """
    today = cfg["extract_date"].isoformat()
    cutoff_dt = datetime.strptime(
        f"{today}T{cutoff_hour:02d}:{cutoff_min:02d}:00", "%Y-%m-%dT%H:%M:%S"
    )
    cutoff_ms = int(cutoff_dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    if cfg["is_admin_console"]:
        must_filter = [
            {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term":  {"Data.role.keyword": "DISTRIBUTOR"}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_field = "Data.syncedUserName.keyword"
    else:
        must_filter = [
            {"term":  {"Data.projectTypeId.keyword": cfg["project_type_id"]}},
            {"term":  {"Data.role.keyword": "DISTRIBUTOR"}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_field = "Data.syncedUserId.keyword"

    query = {
        "size": 0,
        "query": {"bool": {"filter": must_filter}},
        "aggs": {"unique_synced": {"cardinality": {"field": agg_field, "precision_threshold": 3000}}},
    }
    try:
        r = requests.post(
            f"{cfg['es_url']}/{cfg['ES_INDEX_SYNC']}/_search",
            json=query, auth=cfg["es_auth"], verify=False, timeout=30,
        )
        r.raise_for_status()
        count = r.json()["aggregations"]["unique_synced"]["value"]
        log.info(f"  synced by {cutoff_hour:02d}:{cutoff_min:02d} UTC: {count:,}")
        return count
    except Exception as e:
        log.warning(f"  time-cutoff query failed (non-fatal): {e}")
        return None


def _get_synced_keys_by_cutoff(cfg, cutoff_hour=17, cutoff_min=30):
    """
    Return set of CDD keys (lowercased) who synced TODAY before cutoff_hour:cutoff_min UTC.
    Admin console: keys are syncedUserName.lower().
    Project: keys are syncedUserId.lower().
    Returns set, or None if query fails.
    """
    today = cfg["extract_date"].isoformat()
    cutoff_dt = datetime.strptime(
        f"{today}T{cutoff_hour:02d}:{cutoff_min:02d}:00", "%Y-%m-%dT%H:%M:%S"
    )
    cutoff_ms = int(cutoff_dt.replace(tzinfo=timezone.utc).timestamp() * 1000)

    if cfg["is_admin_console"]:
        must_filter = [
            {"term":  {"Data.campaignNumber.keyword": cfg["campaign_number"]}},
            {"term":  {"Data.role.keyword": "DISTRIBUTOR"}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_sources = [{"key": {"terms": {"field": "Data.syncedUserName.keyword"}}}]
    else:
        must_filter = [
            {"term":  {"Data.projectTypeId.keyword": cfg["project_type_id"]}},
            {"term":  {"Data.role.keyword": "DISTRIBUTOR"}},
            {"terms": {"Data.taskDates": [today]}},
            {"range": {"Data.createdTime": {"lte": cutoff_ms}}},
        ]
        agg_sources = [{"key": {"terms": {"field": "Data.syncedUserId.keyword"}}}]

    try:
        buckets = _composite_agg(
            cfg["es_url"], cfg["ES_INDEX_SYNC"], must_filter, agg_sources, cfg["es_auth"]
        )
        synced = {b["key"]["key"].lower() for b in buckets}
        log.info(f"  synced by {cutoff_hour:02d}:{cutoff_min:02d} UTC: {len(synced):,} CDDs")
        return synced
    except Exception as e:
        log.warning(f"  cutoff key query failed (non-fatal): {e}")
        return None


# ── status + build rows ────────────────────────────────────────────────────────

def _sync_status(n, day):
    if n == day:      return "HIGH"
    elif n >= 3:      return "MODERATE"
    elif n >= 1:      return "LOW"
    else:             return "NEVER SYNCED"


def _build_rows(cdds, user_dates, cfg, is_admin):
    day_labels = {}
    start = cfg["campaign_start"]
    for i, dt in enumerate(cfg["CAMPAIGN_DATES"]):
        d = start + __import__("datetime").timedelta(days=i)
        day_labels[dt] = f"Day {i+1} ({d.day} {d.strftime('%b')})"

    all_rows = []
    for key, info in cdds.items():
        dates_synced = user_dates.get(key, set())
        n_days       = len(dates_synced)
        rec = {
            "LGA":             info.get("lga", "") or info.get("province", ""),
            "Health Facility": info.get("facility", "") or info.get("health_area", ""),
            "Username":        info["username"],
            "User ID":         info["user_id"],
            "Days Synced":     n_days,
        }
        for dt, col in day_labels.items():
            rec[col] = "Y" if dt in dates_synced else "N"
        rec["Status"]     = _sync_status(n_days, cfg["DAY"])
        rec["Sync Dates"] = ", ".join(sorted(dates_synced))
        all_rows.append(rec)

    return all_rows, day_labels


# ── Excel writer ───────────────────────────────────────────────────────────────

def _style(cell, fill=None, bold=False, color=None):
    cell.border    = _BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if fill:
        cell.fill = fill
    cell.font = Font(bold=bold, size=9, name="Calibri",
                     color=color if color else "000000")


def _write_df(ws, df, hdr_fill=_HDR_FILL):
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        _style(cell, fill=hdr_fill, bold=True, color="FFFFFF")
    for ri, row in enumerate(df.itertuples(index=False), 2):
        status = getattr(row, "Status", None)
        row_fill = (
            _NEVER_FILL if status == "NEVER SYNCED"
            else _LOW_FILL if status == "LOW"
            else None
        )
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            _style(cell, fill=row_fill)
    # auto-width
    for ci, col in enumerate(df.columns, 1):
        lengths = [len(str(col))] + [len(str(v)) for v in df.iloc[:, ci - 1] if v is not None]
        max_len = max(lengths) if lengths else 10
        ws.column_dimensions[
            __import__("openpyxl.utils", fromlist=["get_column_letter"]).get_column_letter(ci)
        ].width = min(max_len + 2, 30)


def _write_summary(ws, all_rows, cfg):
    lga_stats = defaultdict(lambda: {"total": 0, "HIGH": 0, "MODERATE": 0, "LOW": 0, "NEVER SYNCED": 0})
    for rec in all_rows:
        lg = rec["LGA"] or "Unknown"
        lga_stats[lg]["total"] += 1
        lga_stats[lg][rec["Status"]] += 1

    cols = ["#", "LGA", "Total CDDs", "HIGH", "MODERATE", "LOW", "NEVER SYNCED", "% Never Synced"]
    for ci, h in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _style(cell, fill=_HDR_FILL, bold=True, color="FFFFFF")

    for ri, (lga, s) in enumerate(sorted(lga_stats.items()), 2):
        pct = f"{s['NEVER SYNCED']/s['total']*100:.1f}%" if s["total"] else "-"
        vals = [ri - 1, lga, s["total"], s["HIGH"], s["MODERATE"], s["LOW"], s["NEVER SYNCED"], pct]
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            _style(cell)

    # grand total
    tr = len(lga_stats) + 2
    totals = [
        "", "GRAND TOTAL",
        sum(s["total"] for s in lga_stats.values()),
        sum(s["HIGH"] for s in lga_stats.values()),
        sum(s["MODERATE"] for s in lga_stats.values()),
        sum(s["LOW"] for s in lga_stats.values()),
        sum(s["NEVER SYNCED"] for s in lga_stats.values()),
        "",
    ]
    for ci, val in enumerate(totals, 1):
        cell = ws.cell(row=tr, column=ci, value=val)
        _style(cell, fill=_TOTAL_FILL, bold=True)

    return lga_stats


# ── public entry point ─────────────────────────────────────────────────────────

def run(cfg):
    log.info(f"[cdd_sync] {cfg['state_name']} Day {cfg['DAY']} ...")
    is_admin = cfg["is_admin_console"]

    if is_admin:
        if not cfg["campaign_number"]:
            log.warning("[cdd_sync] is_admin_console=TRUE but campaign_number is empty — skipping")
            return None
        cdds, key_list = _load_staff_admin(cfg)
        user_dates     = _load_syncs_admin(cfg, key_list)
    else:
        if not cfg["project_type_id"]:
            log.warning("[cdd_sync] is_admin_console=FALSE but project_type_id is empty — skipping")
            return None
        cdds, key_list = _load_staff_project(cfg)
        user_dates     = _load_syncs_project(cfg, key_list)

    if not cdds:
        log.warning("[cdd_sync] 0 CDDs found — check campaign_number / project_type_id")
        return None

    all_rows, day_labels = _build_rows(cdds, user_dates, cfg, is_admin)

    COLS = (["LGA", "Health Facility", "Username", "User ID", "Days Synced"]
            + list(day_labels.values()) + ["Status", "Sync Dates"])

    df_all = pd.DataFrame(all_rows).sort_values(["LGA", "Days Synced"])
    for col in COLS:
        if col not in df_all.columns:
            df_all[col] = ""

    # Time-based sync counts
    synced_by_1700 = _count_synced_by_cutoff(cfg, 17, 0)
    synced_by_1730 = _count_synced_by_cutoff(cfg, 17, 30)
    # Keys of CDDs who synced by 17:30 — used for the NOT SYNCED tab
    synced_keys_1730 = _get_synced_keys_by_cutoff(cfg, 17, 30)

    from openpyxl import Workbook as _WB
    wb = _WB()
    wb.remove(wb.active)

    # SUMMARY
    ws_sum = wb.create_sheet("SUMMARY")
    lga_stats = _write_summary(ws_sum, all_rows, cfg)

    # Append time-based stats below grand total in SUMMARY
    total_cdds_count = len(all_rows)
    last_row = ws_sum.max_row + 2
    time_rows = []
    if synced_by_1700 is not None:
        pct = f"{synced_by_1700/total_cdds_count*100:.1f}%" if total_cdds_count else "-"
        time_rows.append(("Synced by 17:00 today (UTC)", synced_by_1700, pct))
    if synced_by_1730 is not None:
        pct = f"{synced_by_1730/total_cdds_count*100:.1f}%" if total_cdds_count else "-"
        time_rows.append(("Synced by 17:30 today (UTC)", synced_by_1730, pct))
    for i, (label, count, pct) in enumerate(time_rows):
        r = last_row + i
        ws_sum.cell(r, 1, label)
        ws_sum.cell(r, 2, count)
        ws_sum.cell(r, 3, pct)
        for ci in range(1, 4):
            _style(ws_sum.cell(r, ci), fill=_TOTAL_FILL, bold=True)

    # Per-LGA
    for lga in sorted(df_all["LGA"].unique()):
        df_lga = df_all[df_all["LGA"] == lga][COLS].copy().reset_index(drop=True)
        df_lga.insert(0, "#", range(1, len(df_lga) + 1))
        safe = str(lga).translate(str.maketrans("/\\*?[]:", "-------"))[:31] or "Unknown"
        ws = wb.create_sheet(safe)
        _write_df(ws, df_lga)

    # NEVER SYNCED
    df_never = df_all[df_all["Status"] == "NEVER SYNCED"][
        ["LGA", "Health Facility", "Username", "User ID"]
    ].reset_index(drop=True)
    df_never.insert(0, "#", range(1, len(df_never) + 1))
    ws_never = wb.create_sheet("NEVER SYNCED")
    _write_df(ws_never, df_never)

    # LOW SYNCED
    df_low = df_all[df_all["Status"] == "LOW"][
        ["LGA", "Health Facility", "Username", "User ID", "Days Synced", "Sync Dates", "Status"]
    ].reset_index(drop=True)
    df_low.insert(0, "#", range(1, len(df_low) + 1))
    ws_low = wb.create_sheet("LOW SYNCED")
    _write_df(ws_low, df_low)

    # NOT SYNCED BY 17:30 — CDDs who had no sync record before 17:30 UTC today
    if synced_keys_1730 is not None:
        lookup_key = "Username" if is_admin else "User ID"
        not_synced = [
            r for r in all_rows
            if r[lookup_key].lower() not in synced_keys_1730
        ]
        df_ns = pd.DataFrame(not_synced)[
            ["LGA", "Health Facility", "Username", "User ID"]
        ].reset_index(drop=True)
        df_ns.insert(0, "#", range(1, len(df_ns) + 1))
        ws_ns = wb.create_sheet("NOT SYNCED BY 17:30")
        _write_df(ws_ns, df_ns)
        log.info(f"  not synced by 17:30 UTC: {len(not_synced):,} CDDs")

    out = cfg["sync_xlsx"]
    wb.save(out)

    total_cdds   = len(all_rows)
    never_count  = sum(1 for r in all_rows if r["Status"] == "NEVER SYNCED")
    log.info(
        f"[cdd_sync] saved -> {out}  "
        f"({total_cdds} CDDs, {never_count} never synced)"
    )
    return out
