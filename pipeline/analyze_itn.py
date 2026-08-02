"""
analyze_itn.py — ITN/LLIN (bed-net) household-based aggregation → performance Excel

Separate module from analyze.py by design: SPAQ/AZM's core computation (is_treated
requires a non-null Data.age; one child = one dose) does not apply to ITN campaigns
(beneficiaryType=HOUSEHOLD, Data.age is always null, one household = N nets based on
household size). Reuses ONLY the generic, drug-agnostic styling primitives from
analyze.py (_style_cell, the fill/border constants, FLAG_COLOR) — analyze.py itself
is never imported for its aggregation logic and is never modified by this file.

Confirmed this session (tenant=chad, campaign CMP-2026-06-03-000312, LLIN_phase2_
tchad2026-final-3):
  - Data.additionalDetails.projectReferenceId (NOT Data.campaignNumber) scopes the
    campaign — this tenant's task docs have no top-level campaignNumber field.
  - Facility tier field is Data.boundaryHierarchy.sppSfd / boundaryHierarchyCode.sppSfd
    (NOT healthFacility/HEALTHFACILITY like SPAQ/AZM tenants).
  - Data.householdId, Data.quantity, Data.memberCount all aggregate correctly
    (cardinality / sum) — validated against live ES this session.

AGGREGATION GRAIN, changed this session: LGA, not health facility. Reasons given
directly by the user:
  1. There are 1,000+ health facilities for this campaign, but only ~40 LGAs — a
     facility-level target was never actually available, and matching 1,000+
     facility names 1:1 is far more fragile than matching ~40 LGA codes.
  2. Real target data for this tenant only exists at LGA grain (confirmed via a
     live DB query this session — see below).
Facility-level detail (records, DQ, low-activity) is STILL scrolled and kept — it
just doesn't carry a target/coverage/status anymore, since no facility-level
target exists. LGA rows carry the real target/coverage/status.

TARGET SOURCE, confirmed this session via live DB + ES cross-check:
  - chad.project_target rows are keyed to chad.project_address.boundary, which for
    boundarytype='LGA' is a CODE like "ADMIN_TC_16_03_ADRE" — NOT the plain LGA
    name ("ADRE") that ES's Data.boundaryHierarchy.district field carries.
  - Confirmed via live ES query: every task doc with boundaryHierarchy.district =
    "ADRE" carries boundaryHierarchyCode.district = "ADMIN_TC_16_03_ADRE" — an
    exact match to the DB's chad.project_target boundary code for that LGA. So
    matching MUST be done on the CODE (boundaryHierarchyCode.district), never on
    the plain name — matching on name would silently produce zero matches.
  - DB query used (chad tenant), beneficiarytype IN ('INDIVIDUAL', 'HOUSEHOLD'):
        SELECT pa.boundary AS lga_code,
               SUM(CASE WHEN pt.beneficiarytype='INDIVIDUAL' THEN pt.targetno ELSE 0 END) AS population_target,
               SUM(CASE WHEN pt.beneficiarytype='HOUSEHOLD'  THEN pt.targetno ELSE 0 END) AS household_target
        FROM chad.project p
        JOIN chad.project_target pt  ON pt.projectid = p.id AND p.isdeleted = false
        JOIN chad.project_address pa ON pa.projectid = p.id AND pa.boundarytype = 'LGA'
        WHERE p.referenceid = '<campaign reference id>'
          AND pt.beneficiarytype IN ('INDIVIDUAL', 'HOUSEHOLD')
        GROUP BY pa.boundary;
  - Net/ITN target (beneficiarytype='PRODUCT') confirmed present too — same query
    with beneficiarytype IN (..., 'PRODUCT') returns real net_target values that
    exactly match an independent cross-check against the campaign's microplan
    Excel (ADRE: 202,850 both ways). Target CSV itself is NOT sourced from the
    microplan (explicitly ruled out) — that was only used once as a cross-check.

DATE SCOPING, added this session — now matches analyze.py's mechanism exactly,
no longer a delta-of-cumulative-snapshots workaround:
  - Confirmed via live ES: Data.taskDates exists on chad's task docs (same field
    name AND format SPAQ already defaults to, e.g. "2026-07-04"), and a real
    GTE/LTE range query against it returns a correct, non-zero, verified count
    (22 records for 2026-07-04) — i.e. it behaves as a proper ES date field, not
    just a keyword string, so the same range-filter pattern analyze.py already
    uses works here without any change.
  - Every run is now scoped to cfg["GTE"]/cfg["LTE"] (config.py already computes
    these for every row, generically — bounds exactly "today" for a daily run, or
    the whole campaign for a --cumulative run), via the SAME field name default
    ("taskDates") and the SAME cfg.get("task_date_field", "taskDates") override
    convention as analyze.py. This means cfg["perf_xlsx"]/itn_history/ snapshots
    are now genuinely TODAY-ONLY (or campaign-wide, in cumulative mode) — no
    longer an always-cumulative-to-date total. report_itn.py sums multiple days'
    snapshots together for the cumulative view, exactly like report.py's
    _load_all_days_perf, instead of subtracting cumulative snapshots.
  - Target division: this campaign's target CSV holds the FULL campaign-length
    target (not a pre-divided daily allocation). For a daily (non-cumulative) run,
    _load_targets_itn's values are divided by the real campaign length in days
    (cfg["campaign_start"] to cfg["campaign_end"], NOT cfg["campaign_days"] —
    config.py defaults that to 4 when a sheet row doesn't set it) before being
    used as each LGA's Status/Coverage denominator — same "Total Campaign Target
    = Daily Target × Campaign Days" relationship analyze.py's own daily mode
    uses. Cumulative runs use the full undivided target directly.
"""
import logging
import os

import pandas as pd
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from pipeline.analyze import (
    _BORDER, _BANNER_FILL, _HDR_FILL, _TOTAL_FILL, _WHITE_FILL, _style_cell,
    _scroll_batches, _fetch_hh_head_map, _fetch_individual_names,
)

urllib3.disable_warnings()
log = logging.getLogger(__name__)

FLAG_COLOR = {
    "HIGH":         "1A7A1A",   # green  >=95%
    "MODERATE":     "E06000",   # orange 70-95%
    "LOW":          "CC0000",   # red    <70%
    "NO TARGET":    "888888",
    "LOW ACTIVITY": "888888",
}

_BATCH = 5000


# ── ES helpers ─────────────────────────────────────────────────────────────────

def _campaign_filter(cfg):
    """
    ITN campaign scoping. Confirmed for tenant 'chad': the task index carries the
    campaign reference at Data.additionalDetails.projectReferenceId, NOT at a
    top-level Data.campaignNumber field (unlike the SPAQ/AZM admin-console tenants).
    """
    if not cfg.get("campaign_number"):
        raise ValueError("campaign_number is required for ITN reporting")
    return {"term": {"Data.additionalDetails.projectReferenceId.keyword": cfg["campaign_number"]}}


def _date_filter(cfg):
    """
    Same mechanism as analyze.py: a GTE/LTE range filter on Data.taskDates
    (override via cfg["task_date_field"], same convention), bounding this run to
    cfg["GTE"]/["LTE"] — which config.py already computes generically for every
    row (today's date for a daily run, the whole campaign span for --cumulative).
    Confirmed via live ES this session: Data.taskDates on chad's task docs is a
    real ES date field (range query returns a correct non-zero count), not just
    a keyword string, so this range filter works the same way it does for SPAQ.
    """
    date_field = cfg.get("task_date_field", "taskDates")
    return {"range": {f"Data.{date_field}": {"gte": cfg["GTE"], "lte": cfg["LTE"]}}}


def _fetch_facility_rows(cfg):
    """
    Scroll-based (NOT a pure ES aggregation) — required for two reasons proven this
    session:
    1. Resolving the household head's real name needs the SAME 2-hop join
       analyze.py uses for SPAQ/AZM (household-member-index -> individual-index).
       A single field-exists check on Data.additionalDetails.familyNameOfIndividual
       was tried and rejected — that field is a device-entered value, sometimes
       incomplete (confirmed: "Dorsso" vs the individual-index's real
       "Hadjidjad Dorsso" for the same person), and the project's own Name
       Resolution Rule says names must never be trusted from additionalDetails.
    2. Nets/population must be counted ONCE per distinct household, not summed
       across every task record. Sync-retry duplicates are real (confirmed: one
       facility had 197 records but only 62 distinct households) — summing
       Data.quantity/Data.memberCount across all of them would triple-count.
       Duplicate volume is still tracked, just as its own DQ metric (dup_records),
       not folded into the coverage numbers.

    Returns facility-level rows (still the natural ES aggregation unit — a task
    doc belongs to one facility), each carrying its LGA name + LGA code so the
    caller can roll them up to LGA grain for target-matching.
    """
    fac_data       = {}   # facility_code -> accumulator dict
    seen_hh_by_fac = {}   # facility_code -> set of already-counted householdIds
    hh_head_ind_map = {}  # householdId -> head's individual clientReferenceId (cached)
    head_name_map   = {}  # individual clientReferenceId -> resolved name (cached)

    _source = [
        "Data.boundaryHierarchy", "Data.boundaryHierarchyCode", "Data.householdId",
        "Data.quantity", "Data.memberCount", "Data.latitude", "Data.longitude",
        "Data.deliveryComments", "Data.additionalDetails.manualCodes",
        "Data.additionalDetails.codesScanned",
    ]
    query = {
        "size": _BATCH,
        "query": {
            "bool": {
                "filter": [
                    _campaign_filter(cfg),
                    _date_filter(cfg),
                    {"term": {"Data.administrationStatus.keyword": "ADMINISTRATION_SUCCESS"}},
                ]
            }
        },
        "_source": _source,
    }

    total_processed = 0
    for batch in _scroll_batches(cfg["es_url"], cfg["ES_INDEX_TASK"], query, cfg["es_auth"], "itn-task"):
        docs = [h["_source"]["Data"] for h in batch]

        # Resolve any new households' head-individual mapping seen in this batch.
        batch_hh_ids = list({d.get("householdId") for d in docs if d.get("householdId")})
        new_hh_ids   = [h for h in batch_hh_ids if h not in hh_head_ind_map]
        if new_hh_ids:
            hh_head_ind_map.update(_fetch_hh_head_map(cfg, new_hh_ids))

        # Resolve any new heads' names seen in this batch.
        batch_head_ids = list({hh_head_ind_map[h] for h in batch_hh_ids if h in hh_head_ind_map})
        new_head_ids   = [i for i in batch_head_ids if i not in head_name_map]
        if new_head_ids:
            head_name_map.update(_fetch_individual_names(cfg, new_head_ids))

        for doc in docs:
            bh  = doc.get("boundaryHierarchy") or {}
            bhc = doc.get("boundaryHierarchyCode") or {}
            province = str(bh.get("province", "") or "").strip()
            lga      = str(bh.get("district", "") or "").strip()
            lga_code = str(bhc.get("district", "") or "").strip()
            fac_name = str(bh.get("sppSfd", "") or "").strip()
            fac_code = str(bhc.get("sppSfd", "") or "").strip()
            if not fac_code:
                continue

            hh_id   = doc.get("householdId") or ""
            lat     = doc.get("latitude")
            lon     = doc.get("longitude")
            add     = doc.get("additionalDetails") or {}
            try:
                manual = int(float(add.get("manualCodes") or 0))
            except (ValueError, TypeError):
                manual = 0
            try:
                scanned = int(float(add.get("codesScanned") or 0))
            except (ValueError, TypeError):
                scanned = 0

            if fac_code not in fac_data:
                fac_data[fac_code] = dict(
                    province=province, lga=lga, lga_code=lga_code,
                    facility_name=fac_name or fac_code,
                    records=0, households_visited=0, nets_distributed=0, population_covered=0,
                    missing_hh_head=0, missing_gps=0,
                    manual_codes=0, scanned_codes=0, missing_codes=0,
                )
                seen_hh_by_fac[fac_code] = set()

            m = fac_data[fac_code]
            m["records"] += 1
            if (lat is None or lat == "") or (lon is None or lon == ""):
                m["missing_gps"] += 1
            m["manual_codes"]  += manual
            m["scanned_codes"] += scanned
            # DQ: a net was distributed but NEITHER scanned NOR manually recorded a
            # code for it — the net has zero barcode documentation, untraceable in
            # inventory. Distinct from the manual-vs-scanned RATIO already tracked.
            if manual == 0 and scanned == 0:
                m["missing_codes"] += 1

            if hh_id and hh_id not in seen_hh_by_fac[fac_code]:
                # First time seeing this household at this facility — count its
                # nets/population/head-name ONCE, deduped against sync-retry repeats.
                seen_hh_by_fac[fac_code].add(hh_id)
                m["households_visited"] += 1
                try:
                    m["nets_distributed"] += int(float(doc.get("quantity") or 0))
                except (ValueError, TypeError):
                    pass
                try:
                    m["population_covered"] += int(float(doc.get("memberCount") or 0))
                except (ValueError, TypeError):
                    pass
                head_ind  = hh_head_ind_map.get(hh_id, "")
                head_name = head_name_map.get(head_ind, "") if head_ind else ""
                if not head_name:
                    m["missing_hh_head"] += 1

        total_processed += len(batch)

    log.info(f"[analyze_itn] {total_processed:,} records processed across {len(fac_data)} facilities")
    return [{"facility_code": code, **m} for code, m in fac_data.items()]


def _aggregate_to_lga(fac_rows):
    """
    Roll facility-level rows up to LGA grain — the grain that actually carries a
    real target (see module docstring). Keyed by lga_code (the DB join key), with
    the plain lga name kept for display.
    """
    lga_data = {}
    for r in fac_rows:
        code = r["lga_code"] or r["lga"] or "UNKNOWN"
        if code not in lga_data:
            lga_data[code] = dict(
                province=r["province"], lga=r["lga"] or code, lga_code=code,
                facilities=0,
                records=0, households_visited=0, nets_distributed=0, population_covered=0,
                missing_hh_head=0, missing_gps=0,
                manual_codes=0, scanned_codes=0, missing_codes=0,
            )
        d = lga_data[code]
        d["facilities"] += 1
        for k in ("records", "households_visited", "nets_distributed", "population_covered",
                  "missing_hh_head", "missing_gps", "manual_codes", "scanned_codes", "missing_codes"):
            d[k] += r[k]
    return list(lga_data.values())


# ── targets ────────────────────────────────────────────────────────────────────

def _load_targets_itn(cfg):
    """
    Reads a pre-built target CSV keyed by LGA CODE (chad.project_target via
    chad.project_address.boundary, boundarytype='LGA' — see module docstring for
    the exact query and the live ES cross-check proving code, not name, is the
    correct join key). Column names are matched case-insensitively against a set
    of known aliases so whatever export format is shared works without another
    round of edits:
      code column:       lga_code | district | boundary | LGA
      population target: population_target | individual_target | TargetPopulation
      household target:   household_target | TargetHouseholds
      net/ITN target:      net_target | product_target | TargetBednets
    Any column not found defaults to 0 (reported, not silently assumed correct).
    """
    csv_path = cfg.get("target_csv", "")
    if not csv_path or not os.path.exists(csv_path):
        log.warning(f"[analyze_itn] target_csv not found: {csv_path} — all targets = 0")
        return {}
    df = pd.read_csv(csv_path)
    cols_lower = {c.lower(): c for c in df.columns}

    def _find(*aliases):
        for a in aliases:
            if a.lower() in cols_lower:
                return cols_lower[a.lower()]
        return None

    code_col = _find("lga_code", "district", "boundary", "lga")
    pop_col  = _find("population_target", "individual_target", "targetpopulation")
    hh_col   = _find("household_target", "targethouseholds")
    net_col  = _find("net_target", "product_target", "targetbednets")

    if not code_col:
        log.warning(f"[analyze_itn] target_csv {csv_path} has no recognisable LGA-code "
                     f"column (tried lga_code/district/boundary/lga) — all targets = 0")
        return {}
    missing = [n for n, c in (("population", pop_col), ("household", hh_col), ("net", net_col)) if not c]
    if missing:
        log.warning(f"[analyze_itn] target_csv {csv_path} missing column(s) for: "
                     f"{', '.join(missing)} — those targets will read 0")

    def _num(row, col):
        if not col:
            return 0
        try:
            return int(float(row.get(col, 0) or 0))
        except (ValueError, TypeError):
            return 0

    tmap = {}
    for _, row in df.iterrows():
        code = str(row.get(code_col, "")).strip()
        if not code:
            continue
        tmap[code] = {
            "household_target":  _num(row, hh_col),
            "population_target": _num(row, pop_col),
            "net_target":        _num(row, net_col),
        }
    log.info(f"[analyze_itn] targets loaded: {len(tmap):,} LGAs")
    return tmap


# ── banding ────────────────────────────────────────────────────────────────────

def _cov_pct(numer, denom):
    return numer / denom * 100 if denom else 0.0


def _band(net_cov_pct, net_target, records):
    # Mirrors analyze.py's _band() priority order exactly: LOW ACTIVITY (too few
    # records to be a meaningful signal) is checked FIRST, before target/coverage.
    if records < 10:
        return "LOW ACTIVITY"
    if net_target == 0:
        return "NO TARGET"
    if net_cov_pct >= 95:
        return "HIGH"
    if net_cov_pct >= 70:
        return "MODERATE"
    return "LOW"


def _finalize_lga_rows(lga_rows, target_map):
    results = []
    for r in sorted(lga_rows, key=lambda x: (x["province"], x["lga"])):
        tgt = target_map.get(r["lga_code"],
                              {"household_target": 0, "population_target": 0, "net_target": 0})
        hh_cov  = _cov_pct(r["households_visited"], tgt["household_target"])
        pop_cov = _cov_pct(r["population_covered"], tgt["population_target"])
        net_cov = _cov_pct(r["nets_distributed"],   tgt["net_target"])
        # DQ signal: records vs distinct households — repeat/duplicate delivery
        # records per household (same spirit as SPAQ's duplicate tracking).
        dup_records = max(0, r["records"] - r["households_visited"])

        results.append({
            **r,
            "household_target":  tgt["household_target"],
            "population_target": tgt["population_target"],
            "net_target":         tgt["net_target"],
            "household_cov":      hh_cov,
            "population_cov":     pop_cov,
            "net_cov":            net_cov,
            "status":             _band(net_cov, tgt["net_target"], r["records"]),
            "dup_records":        dup_records,
        })
    return results


# ── Excel writing ──────────────────────────────────────────────────────────────

HEADERS = [
    "#", "Province", "LGA", "Facilities",
    "Target Households", "Households Visited", "HH Coverage %",
    "Target Population", "Population Covered", "Pop Coverage %",
    "Target ITNs", "Nets Distributed", "ITN Coverage %",
    "Status", "Records", "Duplicate Records",
    "Missing HH Head", "Missing GPS",
    "Manual Codes", "Scanned Codes", "% Scanned", "Missing Codes",
]

# Facility-level detail tab has NO target/coverage/status columns — no facility-
# level target exists (see module docstring), so showing one would be fabricated.
FACILITY_HEADERS = [
    "#", "Province", "LGA", "Health Facility",
    "Records", "Duplicate Records", "Households Visited",
    "Nets Distributed", "Population Covered",
    "Missing HH Head", "Missing GPS",
    "Manual Codes", "Scanned Codes", "% Scanned", "Missing Codes",
]


def _row_values(r, idx):
    total_codes = r["manual_codes"] + r["scanned_codes"]
    pct_scanned = f"{r['scanned_codes']/total_codes*100:.1f}%" if total_codes else "N/A"
    return [
        idx, r["province"], r["lga"], r["facilities"],
        r["household_target"], r["households_visited"], f"{r['household_cov']:.1f}%",
        r["population_target"], r["population_covered"], f"{r['population_cov']:.1f}%",
        r["net_target"], r["nets_distributed"], f"{r['net_cov']:.1f}%",
        r["status"], r["records"], r["dup_records"],
        r["missing_hh_head"], r["missing_gps"],
        r["manual_codes"], r["scanned_codes"], pct_scanned, r["missing_codes"],
    ]


def _facility_row_values(r, idx):
    total_codes = r["manual_codes"] + r["scanned_codes"]
    pct_scanned = f"{r['scanned_codes']/total_codes*100:.1f}%" if total_codes else "N/A"
    return [
        idx, r["province"], r["lga"], r["facility_name"],
        r["records"], r.get("dup_records", 0), r["households_visited"],
        r["nets_distributed"], r["population_covered"],
        r["missing_hh_head"], r["missing_gps"],
        r["manual_codes"], r["scanned_codes"], pct_scanned, r["missing_codes"],
    ]


def _totals_row(rows):
    def s(k): return sum(r[k] for r in rows)
    hh_t, pop_t, net_t = s("household_target"), s("population_target"), s("net_target")
    hh_v, pop_v, net_v = s("households_visited"), s("population_covered"), s("nets_distributed")
    return {
        "province": "", "lga": "GRAND TOTAL", "facilities": s("facilities"),
        "household_target": hh_t, "households_visited": hh_v,
        "household_cov": _cov_pct(hh_v, hh_t),
        "population_target": pop_t, "population_covered": pop_v,
        "population_cov": _cov_pct(pop_v, pop_t),
        "net_target": net_t, "nets_distributed": net_v,
        "net_cov": _cov_pct(net_v, net_t),
        "status": "", "records": s("records"), "dup_records": s("dup_records"),
        "missing_hh_head": s("missing_hh_head"), "missing_gps": s("missing_gps"),
        "manual_codes": s("manual_codes"), "scanned_codes": s("scanned_codes"),
        "missing_codes": s("missing_codes"),
    }


def _write_tab(ws, rows, banner_text):
    ncols = len(HEADERS)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = _BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        _style_cell(cell, fill=_HDR_FILL, bold=True, align="center", color="FFFFFF")

    status_col = HEADERS.index("Status") + 1
    for ri, r in enumerate(rows, 1):
        vals = _row_values(r, ri)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            _style_cell(cell, fill=_WHITE_FILL, align="center")
            if ci == status_col:
                flag_col = FLAG_COLOR.get(str(val), "000000")
                cell.font = Font(bold=True, color=flag_col, size=9, name="Calibri")

    if rows:
        tot = _totals_row(rows)
        tot_row = len(rows) + 3
        tot_vals = _row_values(tot, "")
        for ci, val in enumerate(tot_vals, 1):
            cell = ws.cell(row=tot_row, column=ci, value=val)
            _style_cell(cell, fill=_TOTAL_FILL, bold=True, align="center")

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 12
    for ci in range(5, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


def _write_facility_tab(ws, rows, banner_text):
    ncols = len(FACILITY_HEADERS)
    last_col = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = banner_text
    banner.fill  = _BANNER_FILL
    banner.font  = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20

    for ci, h in enumerate(FACILITY_HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        _style_cell(cell, fill=_HDR_FILL, bold=True, align="center", color="FFFFFF")

    for ri, r in enumerate(sorted(rows, key=lambda x: x["records"]), 1):
        vals = _facility_row_values(r, ri)
        for ci, val in enumerate(vals, 1):
            cell = ws.cell(row=ri + 2, column=ci, value=val)
            _style_cell(cell, fill=_WHITE_FILL, align="center")

    ws.freeze_panes = "E3"
    ws.auto_filter.ref = f"A2:{last_col}2"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 32
    for ci in range(5, ncols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 14


BANDS = ["LOW", "MODERATE", "HIGH", "NO TARGET", "LOW ACTIVITY"]


# ── public entry point ─────────────────────────────────────────────────────────

def run(cfg):
    log.info(f"[analyze_itn] {cfg['state_name']} — streaming ITN task docs ...")

    fac_rows   = _fetch_facility_rows(cfg)
    # Duplicate records per facility, carried into the facility-detail tab (not
    # part of the LGA-level accumulator dict, so computed once here).
    for r in fac_rows:
        r["dup_records"] = max(0, r["records"] - r["households_visited"])

    lga_rows_raw = _aggregate_to_lga(fac_rows)
    target_map   = _load_targets_itn(cfg)

    # Daily runs compare today's activity against a DAILY target, not the full
    # campaign target — same "Total Campaign Target = Daily Target x Campaign
    # Days" relationship analyze.py's own daily mode uses. The target CSV holds
    # the full campaign-length figure, so divide it down here for a daily run;
    # cumulative runs (--cumulative) use the full figure as-is.
    if not cfg.get("cumulative"):
        campaign_start = cfg.get("campaign_start")
        campaign_end   = cfg.get("campaign_end")
        campaign_days  = (campaign_end - campaign_start).days + 1 if campaign_start and campaign_end else None
        if campaign_days:
            target_map = {
                code: {k: round(v / campaign_days) for k, v in t.items()}
                for code, t in target_map.items()
            }

    rows = _finalize_lga_rows(lga_rows_raw, target_map)

    g = _totals_row(rows) if rows else None
    banner_text = (
        f"Target Households: {g['household_target']:,}  |  "
        f"Target Population: {g['population_target']:,}  |  "
        f"Target ITNs: {g['net_target']:,}"
        if g else "No data"
    )

    wb = Workbook()
    wb.remove(wb.active)

    ws_all = wb.create_sheet("ALL LGAS")
    _write_tab(ws_all, rows, banner_text)

    for band in BANDS:
        band_rows = [r for r in rows if r["status"] == band]
        ws = wb.create_sheet(band)
        _write_tab(ws, band_rows, banner_text)

    ws_fac = wb.create_sheet("FACILITY DETAIL")
    _write_facility_tab(ws_fac, fac_rows, banner_text)

    out = cfg["perf_xlsx"]
    wb.save(out)
    log.info(f"[analyze_itn] saved -> {out}  ({len(rows)} LGAs, {len(fac_rows)} facilities)")

    # Day-indexed history snapshot — mirrors SPAQ's per-day file convention
    # (performance_day{N}.xlsx) so report_itn.py's progress chart/cumulative view
    # can read a real historical series, the same way report.py's
    # _load_all_days_perf does. Now that _fetch_facility_rows applies a real
    # GTE/LTE date filter (see _date_filter), each day's snapshot IS genuinely
    # that day's own data — same as SPAQ's per-day files — so report_itn.py sums
    # them together for the cumulative view, exactly like report.py does, rather
    # than needing any delta/subtraction logic.
    # ADAPTED, disclosed: cfg["perf_xlsx"] itself is still unusable for this — it's
    # named from cfg["DAY"], which config.py clamps to cfg["campaign_days"]
    # (defaults to 4 when a Google Sheet row doesn't set it), so every day's run
    # would overwrite the same file rather than accumulating one per day. This
    # writes to a separate, ITN-only history folder keyed by the REAL elapsed
    # day (cfg["campaign_start"] to cfg["extract_date"], both guaranteed present
    # — config.py requires campaign_start/campaign_end on every row), independent
    # of that clamp.
    if cfg.get("campaign_start") and cfg.get("extract_date"):
        elapsed_day = (cfg["extract_date"] - cfg["campaign_start"]).days + 1
        hist_dir = os.path.join(os.path.dirname(out), "itn_history")
        os.makedirs(hist_dir, exist_ok=True)
        hist_path = os.path.join(hist_dir, f"performance_day{elapsed_day}.xlsx")
        wb.save(hist_path)
        log.info(f"[analyze_itn] history snapshot -> {hist_path} (elapsed day {elapsed_day})")

    return out
